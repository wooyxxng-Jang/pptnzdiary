#!/usr/bin/env python3
"""
페퍼톤스_공연DB.xlsx (공연/셋리스트 두 시트, 공연ID로 조인) -> SeedData.json 변환.

출처: 나무위키 '페퍼톤스/단독 공연', '페퍼톤스/외부 공연' 문서 + peppertones.net 공식 사이트.
이 스크립트는 1회성 변환 도구이며, 원본 xlsx가 갱신되면 다시 실행해 SeedData.json을 재생성한다.

사용법:
    python3 convert_seed.py <xlsx_path> <output_json_path>
"""
import json
import re
import sys
import uuid
from datetime import date

import openpyxl

NAMESPACE = uuid.UUID("6b1f2b3a-6c1f-4a1a-9b1a-2f8e6d7c1a10")
SENTINEL_FUTURE_DATE = date(2099, 12, 31)


def parse_date(date_text):
    """원문 날짜 텍스트에서 정렬용 첫 날짜(ISO 8601)를 뽑아낸다."""
    if not date_text:
        return SENTINEL_FUTURE_DATE.isoformat()

    text = str(date_text).strip()

    # "2008.06.20~21", "2013.05.01~05" 등 -> 첫 날짜만 사용
    m = re.match(r"^(\d{4})[.\-](\d{2})[.\-](\d{2})", text)
    if m:
        year, month, day = m.groups()
        if day == "00":  # "2018-06-00"처럼 일자 미상을 00으로 표기한 경우 -> 1일로 보정
            day = "01"
        return f"{year}-{month}-{day}"

    # "2013.06" 처럼 연월만 있는 경우 -> 1일로 보정
    m = re.match(r"^(\d{4})[.\-](\d{2})$", text)
    if m:
        year, month = m.groups()
        return f"{year}-{month}-01"

    return SENTINEL_FUTURE_DATE.isoformat()


def deterministic_id(*parts):
    return str(uuid.uuid5(NAMESPACE, "|".join(parts)))


def load_concerts(ws):
    concerts = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 시트에 시리즈ID(다일차 공연 그룹핑용) 컬럼이 앞에 추가됨 -> 변환에는 사용하지 않음
        _series_id, concert_id, title, date_text, venue, solo, note = row
        if concert_id is None:
            continue
        concerts[concert_id] = {
            "id": deterministic_id("concert", concert_id),
            "title": title,
            "dateText": date_text if date_text else "미정",
            "date": parse_date(date_text),
            "venue": venue if venue else "",
            "tourType": "단독 공연" if solo == "O" else "외부 공연",
            "note": note,
            "sourceNote": "나무위키",
            "setlist": [],
        }
    return concerts


def load_setlist(ws, concerts):
    # 공연 시트가 다일차 공연을 "시리즈ID-회차"(예: C0009-1)로 쪼개 놓은 경우,
    # 셋리스트 시트의 시리즈ID는 여전히 상위 값(C0009)만 가리키므로 회차 컬럼으로 매칭해야 한다.
    series_days = {}
    for concert_id in concerts:
        if "-" in concert_id:
            series_id, _, day = concert_id.rpartition("-")
            series_days.setdefault(series_id, {})[day] = concert_id

    unresolved = []
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        series_id, session, order, song_title, note = row
        if series_id is None:
            continue

        if series_id in concerts:
            concert_id = series_id
        elif series_id in series_days:
            days = series_days[series_id]
            session_key = str(session).strip() if session is not None else None
            if session_key and session_key in days:
                concert_id = days[session_key]
            elif len(days) == 1:
                concert_id = next(iter(days.values()))
            else:
                unresolved.append((row_index, series_id, song_title))
                continue
        else:
            unresolved.append((row_index, series_id, song_title))
            continue

        combined_note = note
        if session:
            combined_note = f"[{session}] {note}" if note else f"[{session}]"

        is_encore = bool(note and "앵콜" in note)

        concerts[concert_id]["setlist"].append({
            "order": int(order),
            "songTitle": song_title,
            "note": combined_note,
            "isEncore": is_encore,
        })

    if unresolved:
        print(f"error: 셋리스트 {len(unresolved)}건을 공연에 연결하지 못했습니다 "
              f"(공연회차 태깅 미완료로 추정). 예시:", file=sys.stderr)
        for row_index, series_id, song_title in unresolved[:10]:
            print(f"  - 셋리스트 시트 {row_index}행: 시리즈ID={series_id}, 곡명={song_title}", file=sys.stderr)
        sys.exit(1)

    for concert in concerts.values():
        concert["setlist"].sort(key=lambda entry: entry["order"])


def main():
    if len(sys.argv) != 3:
        print("usage: convert_seed.py <xlsx_path> <output_json_path>")
        sys.exit(1)

    xlsx_path, output_path = sys.argv[1], sys.argv[2]

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    concerts = load_concerts(wb["공연"])
    load_setlist(wb["셋리스트"], concerts)

    output = list(concerts.values())
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    total_songs = sum(len(c["setlist"]) for c in output)
    print(f"concerts: {len(output)}, setlist entries: {total_songs}")


if __name__ == "__main__":
    main()
